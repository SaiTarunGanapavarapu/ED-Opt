###############################################################################
# function:     saveResultsToExcel()                                          #
# Desctiption:  Function to save results to Excel Sheet                       #
#                                                                             #
# Input:        - results     : Results needed to print to Excel              #
#                 resultsFile : Excel file name to print the results          #                
#                 SheetName   : Excel sheet name to print the results         #
#                                                                             #
# Output:       - Print to write headers and data into Excel sheet            #
#                                                                             #
###############################################################################
import openpyxl
from openpyxl.utils import get_column_letter

def saveResultsToExcel(results, resultsFile, sheetName):
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheetName

    headers = list(results[0].keys())

    # Write headers
    for col, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col)  # 1 → 'A', 27 → 'AA'
        sheet[f"{col_letter}1"] = header

    # Write data rows
    for row, result in enumerate(results, start=2):
        for col, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col)
            value = result[header]
            if isinstance(value, list) or not isinstance(value, (int, float, str)):
                value = str(value)
            sheet[f"{col_letter}{row}"] = value
    
    workbook.save(resultsFile)
    print(f"Results written to {resultsFile}")

